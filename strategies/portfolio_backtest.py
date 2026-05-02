import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment

# ==================== 交易对配置 ====================
# 在这里配置要交易的币种
SYMBOLS = [
    "BTCUSDT",
    # "ETHUSDT",
    # "SOLUSDT",
    # "TRXUSDT",
]

# ==================== 策略参数 ====================
INITIAL_BALANCE = 1000  # 总初始资金 1000 USDT（所有币种共享）
LEVERAGE = 5  # 杠杆倍数
INITIAL_POSITION_PCT = 0.15  # 初次建仓比例（1%）
CONSECUTIVE_THRESHOLD = 8  # 连续次数阈值（4次开始建仓）
POSITION_MULTIPLIER = 1.5  # 加仓倍数
TAKER_FEE = 0.0005  # 吃单手续费 0.05%
# 持仓盈利K线数量：连续 N 根与持仓方向一致的K线后平仓（做空+阴线、做多+阳线）
# =1 表示出现1根盈利K线即平仓；=3 表示连续3根盈利K线后平仓；若不连续则立即平仓
PROFIT_CANDLE_THRESHOLD = 1

# 数据和输出目录
BASE_DIR = Path.home() / "Desktop" / "quant strategy"
DATA_DIR = BASE_DIR / "format data" / "1h datas"  # 数据文件目录
FUNDING_DIR = BASE_DIR / "format data" / "funding_fee"  # 资金费数据目录
OUTPUT_BASE = (
    BASE_DIR / "backtest results"
)  # 回测结果根目录（按日期归类到 YYYYMMDD 子目录）

print("=" * 70)
print("组合策略回测系统")
print("=" * 70)
print(f"交易对: {', '.join(SYMBOLS)}")
print(f"总初始资金: {INITIAL_BALANCE} USDT (所有币种共享)")
print(f"杠杆倍数: {LEVERAGE}x")
print(f"初始仓位: {INITIAL_POSITION_PCT*100}%")
print(f"触发阈值: 连续{CONSECUTIVE_THRESHOLD}次")
print(f"加仓倍数: {POSITION_MULTIPLIER}")
print(f"盈利K线平仓: 连续{PROFIT_CANDLE_THRESHOLD}根盈利K线后平仓（不连续则立即平仓）")
print("=" * 70 + "\n")

# ==================== 加载所有币种数据 ====================
print("正在加载数据...")
all_data = {}

for symbol in SYMBOLS:
    data_file = DATA_DIR / f"{symbol}_1h_all_time.csv"

    if not data_file.exists():
        print(f"✗ {symbol} 数据文件不存在，跳过")
        continue

    df = pd.read_csv(data_file, dtype={"open_time": "int64"})

    # 判断K线涨跌
    df["direction"] = np.where(
        df["close"] > df["open"], 1, np.where(df["close"] < df["open"], -1, 0)
    )

    df["symbol"] = symbol
    all_data[symbol] = df

    print(f"  ✓ {symbol}: {len(df)} 根K线")

if not all_data:
    print("错误: 没有可用的数据文件")
    exit(1)

# ==================== 加载资金费数据 ====================
# 数据源: format data/funding_fee/Funding_Rate_{SYMBOL}_alltime.csv
# 列: symbol, fundingTime, fundingTime_ms, fundingRate, markPrice
# 资金费每8小时结算: UTC 00:00, 08:00, 16:00
funding_dfs = {}
for symbol in all_data.keys():
    funding_path = FUNDING_DIR / f"Funding_Rate_{symbol}_alltime.csv"
    if funding_path.exists():
        fdf = pd.read_csv(funding_path)
        fdf["fundingTime_ms"] = fdf["fundingTime_ms"].astype("int64")
        funding_dfs[symbol] = fdf.set_index("fundingTime_ms")
        print(f"  ✓ {symbol} 资金费: {len(fdf):,} 条")
    else:
        print(f"  ⚠ {symbol} 资金费未找到，将跳过: {funding_path}")

# ==================== 对齐所有币种的时间戳 ====================
print("\n对齐时间戳...")

# 找到所有币种的共同时间范围
min_timestamp = max(df["open_time"].min() for df in all_data.values())
max_timestamp = min(df["open_time"].max() for df in all_data.values())

print(f"  共同时间范围: {min_timestamp} 至 {max_timestamp}")

# 获取基准时间戳列表（使用第一个币种的时间戳）
base_symbol = SYMBOLS[0]
base_timestamps = all_data[base_symbol][
    (all_data[base_symbol]["open_time"] >= min_timestamp)
    & (all_data[base_symbol]["open_time"] <= max_timestamp)
]["open_time"].values

print(f"  基准K线数量: {len(base_timestamps)}")

# 为每个币种创建时间戳索引
for symbol in all_data.keys():
    df = all_data[symbol]
    df_filtered = df[
        (df["open_time"] >= min_timestamp) & (df["open_time"] <= max_timestamp)
    ]

    # 去除重复时间戳（保留第一个）
    df_filtered = df_filtered.drop_duplicates(subset=["open_time"], keep="first")

    df_filtered = df_filtered.set_index("open_time")
    all_data[symbol] = df_filtered

    print(f"  {symbol}: {len(df_filtered)} 根K线（去重后）")


# ==================== 组合策略回测引擎 ====================
class PortfolioStrategy:
    def __init__(self, initial_balance):
        self.balance = initial_balance
        self.initial_balance = initial_balance

        # 每个币种的状态
        self.positions = (
            {}
        )  # {symbol: {'direction': 1/-1, 'size': float, 'entry_price': float}}
        self.consecutive_counts = {}  # {symbol: count}
        self.streak_directions = {}  # {symbol: direction}
        self.consecutive_profit_candles = {}  # {symbol: count} 连续盈利K线数

        for symbol in all_data.keys():
            self.positions[symbol] = None
            self.consecutive_counts[symbol] = 0
            self.streak_directions[symbol] = 0
            self.consecutive_profit_candles[symbol] = 0

        # 记录
        self.trades = []
        self.equity_curve = []
        self.max_balance = initial_balance
        self.max_drawdown = 0

    def _apply_funding_fee(self, symbol, timestamp, current_price):
        """
        若当前时间戳是资金费结算时间且有持仓，则扣收/入账资金费。
        数据源: format data/funding_fee/Funding_Rate_{SYMBOL}_alltime.csv
        资金费公式: fee = 持仓名义价值 × 资金费率
          - 多头: rate > 0 → 付费(余额减少), rate < 0 → 收费(余额增加)
          - 空头: rate > 0 → 收费(余额增加), rate < 0 → 付费(余额减少)
        """
        pos = self.positions.get(symbol)
        if pos is None:
            return
        if symbol not in funding_dfs:
            return
        fdf = funding_dfs[symbol]
        if timestamp not in fdf.index:
            return

        rate = float(fdf.loc[timestamp, "fundingRate"])
        notional = pos["size"] * LEVERAGE  # 名义价值 = 保证金 × 杠杆

        if pos["direction"] == 1:  # 多头
            fee = -notional * rate
        else:  # 空头
            fee = notional * rate

        self.balance += fee
        self.trades.append(
            {
                "timestamp": timestamp,
                "symbol": symbol,
                "action": "FUNDING",
                "direction": "SHORT" if pos["direction"] == -1 else "LONG",
                "price": current_price,
                "size": pos["size"],
                "fee": 0,
                "balance": self.balance,
                "funding_fee": fee,
            }
        )

    def calculate_position_size(self, consecutive_count):
        """计算应有的仓位大小"""
        if consecutive_count < CONSECUTIVE_THRESHOLD:
            return 0

        base_size = self.balance * INITIAL_POSITION_PCT
        n = consecutive_count - CONSECUTIVE_THRESHOLD + 1
        multiplier = POSITION_MULTIPLIER ** (n - 1)

        return base_size * multiplier

    def open_position(self, symbol, price, direction, consecutive_count, timestamp):
        """开仓"""
        size = self.calculate_position_size(consecutive_count)
        # 手续费按持仓价值计算：持仓价值 = 保证金 × 杠杆，fee = 持仓价值 × 0.05%（开平仓各收一次）
        notional = size * LEVERAGE
        fee = notional * TAKER_FEE

        self.positions[symbol] = {
            "direction": -direction,  # 反向
            "size": size,
            "entry_price": price,
        }
        self.consecutive_profit_candles[symbol] = 0

        self.balance -= fee

        self.trades.append(
            {
                "timestamp": timestamp,
                "symbol": symbol,
                "action": "OPEN",
                "direction": "SHORT" if -direction == -1 else "LONG",
                "price": price,
                "size": size,
                "fee": fee,
                "balance": self.balance,
                "consecutive": consecutive_count,
            }
        )

    def add_position(self, symbol, price, consecutive_count, timestamp):
        """加仓"""
        pos = self.positions[symbol]
        new_total_size = self.calculate_position_size(consecutive_count)
        add_size = new_total_size - pos["size"]

        if add_size <= 0:
            return

        # 手续费按持仓价值计算：加仓部分的持仓价值 = add_size × 杠杆
        notional = add_size * LEVERAGE
        fee = notional * TAKER_FEE

        # 更新持仓均价
        total_value = pos["size"] * pos["entry_price"] + add_size * price
        pos["size"] = new_total_size
        pos["entry_price"] = total_value / pos["size"]
        self.balance -= fee

        self.trades.append(
            {
                "timestamp": timestamp,
                "symbol": symbol,
                "action": "ADD",
                "direction": "SHORT" if pos["direction"] == -1 else "LONG",
                "price": price,
                "size": add_size,
                "fee": fee,
                "balance": self.balance,
                "consecutive": consecutive_count,
            }
        )

    def reduce_position(
        self, symbol, price, consecutive_count, timestamp, required_size
    ):
        """减仓：当实际持仓大于所需持仓时，部分平仓使持仓与所需一致"""
        pos = self.positions[symbol]
        reduce_size = pos["size"] - required_size

        if reduce_size <= 0:
            return

        # 计算减仓部分的盈亏
        if pos["direction"] == 1:
            pnl = (
                (price - pos["entry_price"])
                / pos["entry_price"]
                * reduce_size
                * LEVERAGE
            )
        else:
            pnl = (
                (pos["entry_price"] - price)
                / pos["entry_price"]
                * reduce_size
                * LEVERAGE
            )

        # 手续费按减仓部分的持仓价值计算
        notional = reduce_size * LEVERAGE
        fee = notional * TAKER_FEE
        net_pnl = pnl - fee

        self.balance += net_pnl

        # 更新持仓：保留所需部分，entry_price 不变（剩余仓位成本不变）
        pos["size"] = required_size

        self.trades.append(
            {
                "timestamp": timestamp,
                "symbol": symbol,
                "action": "REDUCE",
                "direction": "SHORT" if pos["direction"] == -1 else "LONG",
                "price": price,
                "size": reduce_size,
                "pnl": pnl,
                "fee": fee,
                "net_pnl": net_pnl,
                "balance": self.balance,
                "reason": "持仓大于所需",
                "consecutive": consecutive_count,
            }
        )

    def close_position(self, symbol, price, timestamp, reason="REVERSE"):
        """平仓"""
        pos = self.positions[symbol]

        if pos["direction"] == 1:
            pnl = (
                (price - pos["entry_price"])
                / pos["entry_price"]
                * pos["size"]
                * LEVERAGE
            )
        else:
            pnl = (
                (pos["entry_price"] - price)
                / pos["entry_price"]
                * pos["size"]
                * LEVERAGE
            )

        # 手续费按持仓价值计算：平仓部分的持仓价值 = size × 杠杆
        notional = pos["size"] * LEVERAGE
        fee = notional * TAKER_FEE
        net_pnl = pnl - fee

        self.balance += net_pnl

        self.trades.append(
            {
                "timestamp": timestamp,
                "symbol": symbol,
                "action": "CLOSE",
                "direction": "SHORT" if pos["direction"] == -1 else "LONG",
                "price": price,
                "size": pos["size"],
                "pnl": pnl,
                "fee": fee,
                "net_pnl": net_pnl,
                "balance": self.balance,
                "reason": reason,
            }
        )

        self.positions[symbol] = None
        self.consecutive_profit_candles[symbol] = 0

    def calculate_equity(self, current_prices):
        """计算当前净值（含所有持仓的浮动盈亏）"""
        equity = self.balance

        for symbol, pos in self.positions.items():
            if pos is None:
                continue

            price = current_prices.get(symbol)
            if price is None:
                continue

            if pos["direction"] == 1:
                unrealized_pnl = (
                    (price - pos["entry_price"])
                    / pos["entry_price"]
                    * pos["size"]
                    * LEVERAGE
                )
            else:
                unrealized_pnl = (
                    (pos["entry_price"] - price)
                    / pos["entry_price"]
                    * pos["size"]
                    * LEVERAGE
                )

            equity += unrealized_pnl

        return equity

    def update_symbol(self, symbol, timestamp):
        """更新单个币种的状态"""
        df = all_data[symbol]

        if timestamp not in df.index:
            return

        row = df.loc[timestamp]

        # 处理重复时间戳（取第一行）
        if isinstance(row, pd.DataFrame):
            row = row.iloc[0]

        price = float(row["close"])
        direction = int(row["direction"])

        if direction == 0:
            return

        # 更新连续计数
        if direction == self.streak_directions[symbol]:
            self.consecutive_counts[symbol] += 1
        else:
            self.consecutive_counts[symbol] = 1
            self.streak_directions[symbol] = direction

        consecutive_count = self.consecutive_counts[symbol]
        pos = self.positions[symbol]

        # 策略逻辑
        if pos is None:
            # 无持仓：检查是否开仓
            if consecutive_count >= CONSECUTIVE_THRESHOLD:
                self.open_position(
                    symbol, price, direction, consecutive_count, timestamp
                )

        elif pos["direction"] == direction:
            # 盈利K线：做空+阴线、做多+阳线
            self.consecutive_profit_candles[symbol] += 1
            if self.consecutive_profit_candles[symbol] >= PROFIT_CANDLE_THRESHOLD:
                self.close_position(symbol, price, timestamp, reason="PROFIT_CANDLES")
                if consecutive_count >= CONSECUTIVE_THRESHOLD:
                    self.open_position(
                        symbol, price, direction, consecutive_count, timestamp
                    )
            # else: 继续持有，等待更多连续盈利K线

        else:
            # 亏损K线（趋势继续）：做空+阳线、做多+阴线，不连续则立即平仓
            self.close_position(symbol, price, timestamp, reason="LOSS_CANDLE")
            if consecutive_count >= CONSECUTIVE_THRESHOLD:
                self.open_position(
                    symbol, price, direction, consecutive_count, timestamp
                )

    def run_backtest(self):
        """运行回测"""
        print("\n开始组合回测...")

        for i, timestamp in enumerate(base_timestamps):
            # 获取当前所有币种价格
            current_prices = {}
            for symbol in all_data.keys():
                if timestamp in all_data[symbol].index:
                    close_value = all_data[symbol].loc[timestamp, "close"]
                    # 如果有多个值（重复时间戳），取第一个
                    if isinstance(close_value, pd.Series):
                        current_prices[symbol] = float(close_value.iloc[0])
                    else:
                        current_prices[symbol] = float(close_value)

            # 资金费结算（在策略逻辑前，若当前时间戳是结算时间且有持仓）
            for symbol in all_data.keys():
                if symbol in current_prices:
                    self._apply_funding_fee(symbol, timestamp, current_prices[symbol])

            # 更新每个币种
            for symbol in all_data.keys():
                self.update_symbol(symbol, timestamp)

            # 计算净值
            equity = self.calculate_equity(current_prices)

            # 记录净值
            self.equity_curve.append(
                {"timestamp": timestamp, "balance": self.balance, "equity": equity}
            )

            # 更新最大回撤
            if equity > self.max_balance:
                self.max_balance = equity

            drawdown = (self.max_balance - equity) / self.max_balance
            if drawdown > self.max_drawdown:
                self.max_drawdown = drawdown

            # 爆仓检查
            if equity <= 0:
                print(f"\n⚠️  爆仓！时间: {timestamp}, 净值: {equity:.2f}")
                break

            # 进度显示
            if (i + 1) % 5000 == 0:
                print(
                    f"  进度: {i+1}/{len(base_timestamps)} ({(i+1)/len(base_timestamps)*100:.1f}%)"
                )

        # 平掉所有持仓
        final_timestamp = base_timestamps[-1]
        for symbol, pos in self.positions.items():
            if pos is not None:
                if final_timestamp in all_data[symbol].index:
                    close_value = all_data[symbol].loc[final_timestamp, "close"]
                    if isinstance(close_value, pd.Series):
                        final_price = float(close_value.iloc[0])
                    else:
                        final_price = float(close_value)
                    self.close_position(
                        symbol, final_price, final_timestamp, reason="END"
                    )

        print("回测完成！")


# ==================== 运行回测 ====================
strategy = PortfolioStrategy(INITIAL_BALANCE)
strategy.run_backtest()

# ==================== 统计结果 ====================
print("\n" + "=" * 70)
print("回测结果")
print("=" * 70)

total_trades = len([t for t in strategy.trades if t["action"] == "CLOSE"])
profitable_trades = len(
    [t for t in strategy.trades if t["action"] == "CLOSE" and t["net_pnl"] > 0]
)
win_rate = (profitable_trades / total_trades * 100) if total_trades > 0 else 0
total_pnl = sum([t["net_pnl"] for t in strategy.trades if t["action"] == "CLOSE"])
total_fees = sum([t["fee"] for t in strategy.trades])
total_funding = sum(t.get("funding_fee", 0) for t in strategy.trades)
final_return = (strategy.balance - INITIAL_BALANCE) / INITIAL_BALANCE * 100

print(f"初始资金: {INITIAL_BALANCE:.2f} USDT")
print(f"最终资金: {strategy.balance:.2f} USDT")
print(f"最终净值: {strategy.equity_curve[-1]['equity']:.2f} USDT")
print(f"总收益: {total_pnl:.2f} USDT ({final_return:.2f}%)")
print(f"总手续费: {total_fees:.2f} USDT")
print(f"总资金费: {total_funding:.2f} USDT")
print(f"\n总交易次数: {total_trades}")
print(f"盈利次数: {profitable_trades}")
print(f"胜率: {win_rate:.2f}%")
print(f"最大回撤: {strategy.max_drawdown*100:.2f}%")

# 按币种统计
print(f"\n各币种交易次数:")
for symbol in SYMBOLS:
    count = len(
        [t for t in strategy.trades if t["symbol"] == symbol and t["action"] == "CLOSE"]
    )
    print(f"  {symbol}: {count} 次")

print("=" * 70)

# ==================== 保存结果（按日期归类，参考 arc_top_backtest） ====================
trades_df = pd.DataFrame(strategy.trades)
equity_df = pd.DataFrame(strategy.equity_curve)

date_str = datetime.now().strftime("%Y%m%d")
output_dir = OUTPUT_BASE / date_str
output_dir.mkdir(parents=True, exist_ok=True)
ts = datetime.now().strftime("%Y%m%d_%H%M%S")
syms = "_".join(SYMBOLS)

trades_path = output_dir / f"portfolio_trades_{syms}_{ts}.csv"
equity_path = output_dir / f"portfolio_equity_{syms}_{ts}.csv"

trades_df.to_csv(trades_path, index=False, encoding="utf-8-sig")
equity_df.to_csv(equity_path, index=False, encoding="utf-8-sig")

print(f"\n详细数据已保存:")
print(f"  - {trades_path.name}")
print(f"  - {equity_path.name}")

# ==================== 创建Excel图表 ====================
print("\n正在创建Excel图表...")

# 转换时间戳
equity_df["date"] = pd.to_datetime(equity_df["timestamp"], unit="ms").dt.strftime(
    "%Y-%m-%d %H:%M"
)
output_df = equity_df[["date", "balance", "equity"]].copy()
output_df.columns = ["时间", "账户余额", "账户净值"]

# 创建Excel
excel_path = output_dir / f"portfolio_equity_chart_{syms}_{ts}.xlsx"

with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
    output_df.to_excel(writer, sheet_name="净值数据", index=False)

# 添加图表
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

wb = load_workbook(excel_path)
ws = wb["净值数据"]

# 设置列宽
ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 15

# 表头样式
header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

for col in ["A", "B", "C"]:
    cell = ws[f"{col}1"]
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 数字格式
for row in range(2, len(output_df) + 2):
    ws.cell(row=row, column=2).number_format = "#,##0.00"
    ws.cell(row=row, column=3).number_format = "#,##0.00"

# 创建折线图
chart = LineChart()
chart.title = "组合策略净值曲线"
chart.style = 10
chart.y_axis.title = "净值 (USDT)"
chart.x_axis.title = "时间"
chart.height = 15
chart.width = 30

# 添加数据
max_row = len(output_df) + 1
data_equity = Reference(ws, min_col=3, min_row=1, max_row=max_row)
data_balance = Reference(ws, min_col=2, min_row=1, max_row=max_row)

chart.add_data(data_equity, titles_from_data=True)
chart.add_data(data_balance, titles_from_data=True)

# 设置线条颜色
chart.series[0].graphicalProperties.line.solidFill = "4472C4"  # 净值-蓝色
chart.series[0].graphicalProperties.line.width = 20000
chart.series[1].graphicalProperties.line.solidFill = "ED7D31"  # 余额-橙色
chart.series[1].graphicalProperties.line.width = 20000

chart.legend.position = "b"

ws.add_chart(chart, "E2")

# 添加统计
stats_row = len(output_df) + 5
stats_data = [
    ["回测统计", ""],
    ["初始资金", f"{INITIAL_BALANCE:.2f} USDT"],
    ["最终余额", f"{strategy.balance:.2f} USDT"],
    ["最终净值", f"{output_df['账户净值'].iloc[-1]:.2f} USDT"],
    ["收益率", f"{final_return:.2f}%"],
    ["最大回撤", f"{strategy.max_drawdown*100:.2f}%"],
    ["总交易次数", f"{total_trades}"],
    ["胜率", f"{win_rate:.2f}%"],
]

for i, (label, value) in enumerate(stats_data):
    row = stats_row + i
    ws[f"E{row}"] = label
    ws[f"F{row}"] = value

    if i == 0:
        ws[f"E{row}"].font = Font(name="Arial", size=12, bold=True)
    else:
        ws[f"E{row}"].font = Font(name="Arial", size=10, bold=True)
        ws[f"F{row}"].font = Font(name="Arial", size=10)

ws.freeze_panes = "A2"

wb.save(excel_path)

print(f"✓ Excel图表已生成: {excel_path.name}")
print(f"\n所有结果保存在: {output_dir}")
